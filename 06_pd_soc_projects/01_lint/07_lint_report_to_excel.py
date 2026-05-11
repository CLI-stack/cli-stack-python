"""
Project: Lint Report to Excel Exporter
Tool context: Spyglass Lint / Questa Lint
What it does: Converts a lint report to a color-coded Excel spreadsheet.
Fatal rows = red, Error = orange, Warning = yellow, Info = blue.
Adds a summary sheet with charts. Perfect for sharing with leads or customers.

Install: pip install openpyxl
Usage:
    python 07_lint_report_to_excel.py --report lint.rpt --output lint_report.xlsx
"""

import re
import argparse
from collections import Counter

SAMPLE_REPORT = """\
Fatal|W_COMB_LOOP|fifo_ctrl|rtl/fifo_ctrl.v:134|Combinational loop detected
Error|W_NET_NO_LOAD|cpu_core|rtl/cpu_core.v:245|Net 'int_bus' has no load
Error|W_NET_NO_LOAD|cpu_core|rtl/cpu_core.v:312|Net 'dbg_out' has no load
Error|W_MUX_SEL_UNDRIVEN|apb_bridge|rtl/apb_bridge.v:67|Mux select undriven
Warning|W_REGS_NO_RESET|alu_unit|rtl/alu_unit.v:88|Register has no reset
Warning|W_REGS_NO_RESET|dma_ctrl|rtl/dma_ctrl.v:120|Register has no reset
Warning|W_PORT_UNCONNECTED|soc_top|rtl/soc_top.v:200|Port unconnected
Info|W_CONST_DRIVER|dma_ctrl|rtl/dma_ctrl.v:55|Signal driven by constant
"""

# Row fill colors per severity (Excel ARGB format)
SEV_COLORS = {
    "Fatal":   "FFFF0000",   # red
    "Error":   "FFFF8C00",   # dark orange
    "Warning": "FFFFD700",   # gold
    "Info":    "FFB0C4DE",   # light steel blue
}

SEV_RANK = {"Fatal": 0, "Error": 1, "Warning": 2, "Info": 3}


def parse_violations(text):
    pattern = re.compile(r"^(Fatal|Error|Warning|Info)\|(\S+)\|(\S+)\|(.+):(\d+)\|(.+)$")
    violations = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        m = pattern.match(line)
        if m:
            violations.append({
                "Severity": m.group(1), "Rule":    m.group(2),
                "Module":   m.group(3), "File":    m.group(4),
                "Line":     int(m.group(5)), "Message": m.group(6),
            })
    return sorted(violations, key=lambda v: SEV_RANK.get(v["Severity"], 9))


def export_excel(violations, output):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()

    # ── Sheet 1: All violations ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Violations"

    headers = ["Severity", "Rule", "Module", "File", "Line", "Message"]
    header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, v in enumerate(violations, 2):
        fill_color = SEV_COLORS.get(v["Severity"], "FFFFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=v[key])
            cell.fill = fill
            if v["Severity"] in ("Fatal", "Error"):
                cell.font = Font(bold=True)

    # Set column widths
    col_widths = [10, 28, 18, 35, 7, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    sev_counts  = Counter(v["Severity"] for v in violations)
    rule_counts = Counter(v["Rule"]     for v in violations)
    mod_counts  = Counter(v["Module"]   for v in violations)

    ws2["A1"] = "LINT REPORT SUMMARY"
    ws2["A1"].font = Font(bold=True, size=14)

    ws2["A3"] = "Severity"
    ws2["B3"] = "Count"
    for i, sev in enumerate(["Fatal", "Error", "Warning", "Info"], 4):
        ws2[f"A{i}"] = sev
        ws2[f"B{i}"] = sev_counts.get(sev, 0)
        fill = PatternFill(start_color=SEV_COLORS[sev], end_color=SEV_COLORS[sev], fill_type="solid")
        ws2[f"A{i}"].fill = fill
        ws2[f"B{i}"].fill = fill

    ws2["D3"] = "Top Rules"
    ws2["E3"] = "Count"
    for i, (rule, cnt) in enumerate(rule_counts.most_common(10), 4):
        ws2[f"D{i}"] = rule
        ws2[f"E{i}"] = cnt

    ws2["G3"] = "Top Modules"
    ws2["H3"] = "Count"
    for i, (mod, cnt) in enumerate(mod_counts.most_common(10), 4):
        ws2[f"G{i}"] = mod
        ws2[f"H{i}"] = cnt

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["D"].width = 28
    ws2.column_dimensions["G"].width = 20

    wb.save(output)
    print(f"Excel report saved: {output}  ({len(violations)} violations)")


# ── Main ─────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="Export lint report to color-coded Excel")
parser.add_argument("--report", help="Lint report file")
parser.add_argument("--output", default="lint_report.xlsx")
args = parser.parse_args()

text = open(args.report).read() if args.report else SAMPLE_REPORT
if not args.report:
    print("No report file given — using sample data.\n")

violations = parse_violations(text)
print(f"Parsed {len(violations)} violations. Generating Excel...")
export_excel(violations, args.output)
