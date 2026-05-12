"""
Project: Excel Report Generator
What it does: Creates a formatted Excel spreadsheet report with multiple sheets,
color-coded cells, summary statistics, and a chart.
Useful for generating automated business reports.

Install: pip install openpyxl
Run: python 08_excel_report_generator.py
Run: python 08_excel_report_generator.py --output my_report.xlsx
"""

import argparse
import os

try:
    from openpyxl import Workbook                     # create Excel workbooks
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side    # cell formatting
    )
    from openpyxl.utils import get_column_letter      # convert col number to letter
    from openpyxl.chart import BarChart, Reference    # charts

    # ── Sample sales data ────────────────────────────────────────────────────
    SALES_DATA = [
        # (month, product, units, revenue, target)
        ("January",  "Laptop",  45,  54000, 50000),
        ("January",  "Phone",   120, 36000, 40000),
        ("January",  "Tablet",  60,  18000, 20000),
        ("February", "Laptop",  52,  62400, 50000),
        ("February", "Phone",   130, 39000, 40000),
        ("February", "Tablet",  55,  16500, 20000),
        ("March",    "Laptop",  61,  73200, 60000),
        ("March",    "Phone",   145, 43500, 45000),
        ("March",    "Tablet",  72,  21600, 22000),
    ]

    def make_fill(hex_color):
        """Create a cell background fill using a hex color code like 'FF0000'."""
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def make_border():
        """Create a thin border around a cell."""
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def generate_report(output_path):
        """Build the Excel workbook with multiple sheets."""
        wb = Workbook()  # create a new empty workbook

        # ── Sheet 1: Raw Sales Data ───────────────────────────────────────────
        ws1 = wb.active         # the first sheet is created automatically
        ws1.title = "Sales Data"

        # Define column headers
        headers = ["Month", "Product", "Units Sold", "Revenue ($)", "Target ($)", "Status"]

        # Write header row with formatting
        header_fill = make_fill("2C3E50")   # dark blue background
        header_font = Font(color="FFFFFF", bold=True)  # white bold text

        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows with conditional formatting
        for row_idx, (month, product, units, revenue, target) in enumerate(SALES_DATA, start=2):
            # Determine if this row met its target (for color coding)
            met_target = revenue >= target
            row_fill   = make_fill("D4EDDA") if met_target else make_fill("F8D7DA")
            # Green if target met, red if not

            status = "✓ MET" if met_target else "✗ MISSED"

            # Write each cell in this row
            row_data = [month, product, units, revenue, target, status]
            for col, value in enumerate(row_data, start=1):
                cell            = ws1.cell(row=row_idx, column=col, value=value)
                cell.fill       = row_fill
                cell.border     = make_border()
                cell.alignment  = Alignment(horizontal="center")

                # Bold the status column
                if col == 6:
                    cell.font = Font(bold=True,
                                     color="155724" if met_target else "721C24")

        # Set column widths for readability
        column_widths = [12, 10, 12, 14, 12, 12]
        for i, width in enumerate(column_widths, start=1):
            ws1.column_dimensions[get_column_letter(i)].width = width

        # ── Sheet 2: Summary by Month ─────────────────────────────────────────
        ws2 = wb.create_sheet("Monthly Summary")

        ws2.append(["Month", "Total Revenue", "Total Target", "Variance", "Achievement %"])

        # Group data by month and calculate totals
        from collections import defaultdict
        monthly = defaultdict(lambda: {"revenue": 0, "target": 0})

        for month, _, _, revenue, target in SALES_DATA:
            monthly[month]["revenue"] += revenue
            monthly[month]["target"]  += target

        # Write summary rows
        header_fill2 = make_fill("1A252F")
        for col in range(1, 6):
            cell = ws2.cell(row=1, column=col)
            cell.fill = header_fill2
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (month, data) in enumerate(monthly.items(), start=2):
            rev      = data["revenue"]
            tgt      = data["target"]
            variance = rev - tgt                    # positive = beat target
            pct      = rev / tgt * 100 if tgt else 0  # achievement percentage

            fill_color = "D4EDDA" if variance >= 0 else "F8D7DA"
            row_fill   = make_fill(fill_color)

            for col, value in enumerate([month, rev, tgt, variance, round(pct, 1)], start=1):
                cell           = ws2.cell(row=row_idx, column=col, value=value)
                cell.fill      = row_fill
                cell.border    = make_border()
                cell.alignment = Alignment(horizontal="center")

        ws2.column_dimensions["A"].width = 12
        for col in ["B", "C", "D", "E"]:
            ws2.column_dimensions[col].width = 16

        # ── Sheet 3: Bar Chart ─────────────────────────────────────────────────
        ws3 = wb.create_sheet("Chart")

        # Add data for the chart
        ws3.append(["Month", "Revenue", "Target"])
        for month, data in monthly.items():
            ws3.append([month, data["revenue"], data["target"]])

        # Create a bar chart
        chart = BarChart()
        chart.type    = "col"             # "col" = vertical bars
        chart.title   = "Revenue vs Target by Month"
        chart.y_axis.title = "Amount ($)"
        chart.x_axis.title = "Month"

        # Define which data to use for the chart (rows 1-4, columns 2-3)
        data_ref   = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=4)
        cats_ref   = Reference(ws3, min_col=1, min_row=2, max_row=4)
        chart.add_data(data_ref, titles_from_data=True)  # first row = legend
        chart.set_categories(cats_ref)

        ws3.add_chart(chart, "E2")  # place chart starting at cell E2

        # ── Save the workbook ──────────────────────────────────────────────────
        wb.save(output_path)
        print(f"Excel report saved: {output_path}")
        print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

    # ── Main ─────────────────────────────────────────────────────────────────
    parser = argparse.ArgumentParser(description="Generate formatted Excel report")
    parser.add_argument("--output", default="sales_report.xlsx")
    args = parser.parse_args()

    print("=== Excel Report Generator ===\n")
    generate_report(args.output)

    # Clean up
    if os.path.exists(args.output):
        os.remove(args.output)
        print("(Report file cleaned up for demo)")

except ImportError:
    print("Install: pip install openpyxl")
